from openpyxl import load_workbook
import openpyxl
import os
import pandas as pd
from src.utils import data_cleaner
import re
 
 
def Find(string):
 
    # Use a regular expression pattern to match URLs
    pattern = re.compile(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+')
    links = re.findall(pattern, string)
    
    # Return the first link if any are found, or None if not
    return links[0] if links else None

def extract_report_number(url):
    if url is None:
        return None
    report_number = re.search("(reportno|report_no)=(\d+)", url)
    if report_number is None:
        return None
    return report_number.group(2)

test_data_dir  = os.path.join("artifacts","new_test_data_2")
test_data_files_name =  os.listdir(test_data_dir)

import pandas as pd
from openpyxl import load_workbook

for file_name in test_data_files_name[:]:
    if file_name != "UNIQUE BR.xlsx":
        continue
    print()
    print(f" =========== File Name : {file_name} ============")
    
    # if file_name != "MSURESH.xlsx":
    #     continue

    df = pd.read_excel(os.path.join(test_data_dir,file_name), None)
    sheet_names = list(df.keys())
    cur_sheet = df[sheet_names[0]]
    columns_name = list(cur_sheet.columns)

    wb = openpyxl.load_workbook(os.path.join(test_data_dir,file_name))
    ws = wb[sheet_names[0]]

    cols_link = []

    for cur_col in range(1,len(columns_name)+1):
        
        if ws.cell(row = 12, column = cur_col).value is not None:
            cur_link = ws.cell(row = 12, column = cur_col).hyperlink

            if cur_link is not None:
                if cur_link.target is not None:
                    cols_link.append((columns_name[cur_col-1],cur_col))

            else:
                if type(ws.cell(row = 12, column = cur_col).value) == str:
                    if Find(ws.cell(row = 12, column = cur_col).value) is not None:
                        cols_link.append((columns_name[cur_col-1],cur_col))

    cur_sheet = df[sheet_names[0]]
    df,correct_row_idx = data_cleaner.correct_df_headers(cur_sheet)
    print(len(df))
    df_link = pd.DataFrame()

    print(f"==== {cols_link} ====")

    total_link_columns = len(cols_link)

    if total_link_columns == 0:
        continue

    for j in range(len(cols_link)):
        df_link[df.columns[cols_link[j][1]-1] + "_link"] = [None]*len(df)
    
    # Initializing report number column
    df_link['report_no'] = [None]*len(df)

    t = 0

    for i in range(correct_row_idx+3,ws.max_row+1):
        
        for j in range(len(cols_link)):

            cur_cell = ws.cell(row=i,column=cols_link[j][1])
            
            # If completely empty simply return None
            if cur_cell.value is None and cur_cell.hyperlink is None:
                try:
                    df_link[df.columns[cols_link[j][1]-1] + "_link"].iloc[t] = None
                except:
                    pass
                continue

            # If hyperlink is not None
            cur_hyperlink_value = cur_cell.hyperlink

            if cur_hyperlink_value is not None:
                
                # If hyperlink is not given
                if cur_hyperlink_value.target is not None:
                    extracted_link = cur_hyperlink_value.target
                    df_link[df.columns[cols_link[j][1]-1] + "_link"].iloc[t] = cur_hyperlink_value.target

            # If hyperlink is None or hyperlink.target is None
            else:
                extracted_link = Find(cur_cell.value)
            
            df_link[df.columns[cols_link[j][1]-1] + "_link"].iloc[t] = extracted_link
            if df_link['report_no'].iloc[t] is None:
                df_link['report_no'].iloc[t] = extract_report_number(extracted_link)
            
        t+=1

    print("+==========================+")

    

    # Drop column if entirely empty
    df_link.dropna(axis=1, how='all', inplace=True)
    
    print(len(df_link) == len(df))

    df = pd.concat([df,df_link],axis=1)
    print(df.tail())