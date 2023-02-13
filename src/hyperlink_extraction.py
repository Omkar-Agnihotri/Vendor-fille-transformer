from openpyxl import load_workbook
import openpyxl
import os
import pandas as pd
import re
import urllib.parse
from src.utils import hyperlink_extraction_utils

class HyperlinkExtractor:
    def __init__(self, ws, correct_row_idx, df):
        self.ws = ws
        self.correct_row_idx = correct_row_idx
        self.df = df
        self.columns_name = list(df.columns)
        self.cols_link = hyperlink_extraction_utils.get_hyperlink_columns(df, ws, self.columns_name)
        self.total_link_columns = len(self.cols_link)
        self.df_link = pd.DataFrame()
        self.new_columns = []
        
    def add_hyperlink_columns(self):
        if self.total_link_columns == 0:
            return self.df, []
        for j in range(len(self.cols_link)):
            self.df_link[self.df.columns[self.cols_link[j][1]-1] + "_link"] = [None]*len(self.df)
        self.df_link['report_no'] = [None]*len(self.df)
        self.df_link = self.iteratively_extract_link_and_report_number(0)
        self.df_link.dropna(axis=1, how='all', inplace=True)
        self.df = pd.concat([self.df, self.df_link], axis=1)
        self.new_columns = list(self.df_link.columns)
        return self.df, self.new_columns
    
    def iteratively_extract_link_and_report_number(self, t):
        for i in range(self.correct_row_idx+3, self.ws.max_row+1):
            for j in range(len(self.cols_link)):
                cur_cell = self.ws.cell(row=i, column=self.cols_link[j][1])
                extracted_link = self.extract_link_from_current_cell(cur_cell)
                self.df_link[self.df.columns[self.cols_link[j][1]-1] + "_link"].iloc[t] = extracted_link
                if self.df_link['report_no'].iloc[t] is None:
                    self.df_link['report_no'].iloc[t] = hyperlink_extraction_utils.extract_report_number(extracted_link)
            t += 1
        return self.df_link
    
    def extract_link_from_current_cell(self, cur_cell):
        if cur_cell.value is None and cur_cell.hyperlink is None:
            return None
        cur_hyperlink_value = cur_cell.hyperlink
        if cur_hyperlink_value is not None and cur_hyperlink_value.target is not None:
            return cur_hyperlink_value.target
        return hyperlink_extraction_utils.Find(cur_cell.value)